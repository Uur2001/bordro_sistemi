"""
Bordro Excel Export Modülü
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from decimal import Decimal


def format_para(sayi):
    """Sayıyı Türk Lirası formatına çevir"""
    if sayi is None:
        return "0,00"
    if isinstance(sayi, Decimal):
        sayi = float(sayi)
    return f"{sayi:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def create_aylik_bordro_excel(sonuc, calisan_adi=None, ay=None, yil=None):
    """Aylık bordro için Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bordro"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    section_font = Font(bold=True, size=10)
    number_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    row = 1

    donem = sonuc.get('donem', {})
    ay_adi = donem.get('ay_adi', {}).get('ad', 'Ocak') if isinstance(donem.get('ay_adi'), dict) else 'Ocak'
    yil_val = donem.get('yil', 2026)

    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"BORDRO - {ay_adi} {yil_val}"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1

    if calisan_adi:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Çalışan: {calisan_adi}"
        ws[f'A{row}'].font = Font(size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

    row += 1

    ws[f'A{row}'] = "KAZANÇLAR"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    headers = ['Kazanç', 'Brüt', 'SGK+İşs.+BES', 'GV', 'DV', 'Net']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    temel = sonuc.get('temel_ucret', {})
    brut_kazanc = sonuc.get('brut_kazanclar', {})
    sgk = sonuc.get('sgk', {}).get('primler', {})
    gv = sonuc.get('gelir_vergisi', {})
    dv = sonuc.get('damga_vergisi', {})
    net = sonuc.get('net_ucret', 0)

    brut = temel.get('bu_ayki_temel_ucret', 0) or brut_kazanc.get('temel_ucret', 0)
    sgk_toplam = sgk.get('isci_toplam', 0)
    gv_val = gv.get('odenecek', 0)
    dv_val = dv.get('odenecek_dv', 0)

    data = [
        'Temel Ücret',
        format_para(brut),
        format_para(sgk_toplam),
        format_para(gv_val),
        format_para(dv_val),
        format_para(net)
    ]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        if col > 1:
            cell.alignment = Alignment(horizontal='right')
    row += 2

    ws[f'A{row}'] = "SOSYAL GÜVENLİK PRİMLERİ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    sgk_headers = ['Prim Türü', 'Oran (İşçi)', 'Tutar (İşçi)', 'Oran (İşveren)', 'Tutar (İşveren)', '']
    for col, header in enumerate(sgk_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    oranlar = sgk.get('oranlar', {})

    ws.cell(row=row, column=1, value="SGK Primi").border = thin_border
    ws.cell(row=row, column=2, value=f"%{oranlar.get('sgk_isci', 14):.2f}").border = thin_border
    ws.cell(row=row, column=3, value=format_para(sgk.get('isci_sgk', 0))).border = thin_border
    ws.cell(row=row, column=4, value=f"%{oranlar.get('sgk_isveren', 19.5):.2f}").border = thin_border
    ws.cell(row=row, column=5, value=format_para(sgk.get('isveren_sgk', 0))).border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="İşsizlik Sigortası").border = thin_border
    ws.cell(row=row, column=2, value=f"%{oranlar.get('issizlik_isci', 1):.2f}").border = thin_border
    ws.cell(row=row, column=3, value=format_para(sgk.get('isci_issizlik', 0))).border = thin_border
    ws.cell(row=row, column=4, value=f"%{oranlar.get('issizlik_isveren', 2):.2f}").border = thin_border
    ws.cell(row=row, column=5, value=format_para(sgk.get('isveren_issizlik', 0))).border = thin_border
    row += 1

    bes = sgk.get('isci_bes', 0)
    if bes > 0:
        ws.cell(row=row, column=1, value="BES Kesintisi").border = thin_border
        ws.cell(row=row, column=2, value="%3.00").border = thin_border
        ws.cell(row=row, column=3, value=format_para(bes)).border = thin_border
        ws.cell(row=row, column=4, value="-").border = thin_border
        ws.cell(row=row, column=5, value="-").border = thin_border
        row += 1

    hazine = sgk.get('hazine_yardimi', 0)
    if hazine > 0:
        ws.cell(row=row, column=1, value="Hazine Yardımı (-)").border = thin_border
        ws.cell(row=row, column=2, value="-").border = thin_border
        ws.cell(row=row, column=3, value="-").border = thin_border
        ws.cell(row=row, column=4, value=f"%{oranlar.get('hazine_yardimi', 2):.2f}").border = thin_border
        ws.cell(row=row, column=5, value=format_para(hazine)).border = thin_border
        row += 1

    ws.cell(row=row, column=1, value="TOPLAM").border = thin_border
    ws[f'A{row}'].font = Font(bold=True)
    ws.cell(row=row, column=2, value="").border = thin_border
    ws.cell(row=row, column=3, value=format_para(sgk_toplam)).border = thin_border
    ws[f'C{row}'].font = Font(bold=True)
    ws.cell(row=row, column=4, value="").border = thin_border
    ws.cell(row=row, column=5, value=format_para(sgk.get('isveren_toplam', 0))).border = thin_border
    ws[f'E{row}'].font = Font(bold=True)
    row += 2

    ws[f'A{row}'] = "VERGİLER"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    ws.cell(row=row, column=1, value="Gelir Vergisi Matrahı").border = thin_border
    ws.cell(row=row, column=2, value=format_para(gv.get('matrah', 0))).border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Hesaplanan GV").border = thin_border
    ws.cell(row=row, column=2, value=format_para(gv.get('hesaplanan', 0))).border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Asgari Ücret GV İstisnası").border = thin_border
    ws.cell(row=row, column=2, value=format_para(gv.get('istisna', {}).get('istisna_vergi', 0))).border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Ödenecek GV").border = thin_border
    ws.cell(row=row, column=2, value=format_para(gv_val)).border = thin_border
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

    # DV
    ws.cell(row=row, column=1, value="Damga Vergisi Matrahı").border = thin_border
    ws.cell(row=row, column=2, value=format_para(dv.get('brut_kazanc', 0))).border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Ödenecek DV").border = thin_border
    ws.cell(row=row, column=2, value=format_para(dv_val)).border = thin_border
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 2

    ws[f'A{row}'] = "ÖZET"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    maliyet = sonuc.get('isveren_maliyeti', {})

    ozet_data = [
        ('Brüt Kazanç', brut_kazanc.get('toplam_brut', 0)),
        ('SGK + İşsizlik + BES (İşçi)', sgk_toplam),
        ('Gelir Vergisi', gv_val),
        ('Damga Vergisi', dv_val),
        ('Toplam Kesinti', sonuc.get('kesintiler', {}).get('toplam', 0)),
        ('NET ÜCRET', net),
        ('', ''),
        ('İşveren SGK + İşsizlik', maliyet.get('isveren_prim_toplami', 0)),
        ('Hazine Yardımı (-)', maliyet.get('hazine_yardimi', 0)),
        ('TOPLAM MALİYET', maliyet.get('toplam_maliyet', 0)),
    ]

    for label, val in ozet_data:
        ws.cell(row=row, column=1, value=label).border = thin_border
        if val != '':
            ws.cell(row=row, column=2, value=format_para(val)).border = thin_border
        if label in ['NET ÜCRET', 'TOPLAM MALİYET']:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_yillik_bordro_excel(sonuc, calisan_adi=None, yil=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Yıllık Bordro"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"YILLIK BORDRO - {yil or 2026}"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1

    if calisan_adi:
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"Çalışan: {calisan_adi}"
        row += 1

    row += 1
    widths = {'A': 12, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    aylik_sonuclar = sonuc.get('aylik_sonuclar', []) if isinstance(sonuc, dict) else []

    if aylik_sonuclar and isinstance(aylik_sonuclar, list) and len(aylik_sonuclar) > 0:
        ws[f'A{row}'] = "AYLIK DETAYLAR"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        aylik_basliklar = ['Ay', 'Brüt', 'SGK (İşçi)', 'İşsizlik', 'GV', 'Kesinti', 'Net', 'İşv. Maliyet']
        for col, header in enumerate(aylik_basliklar, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        for ay_sonuc in aylik_sonuclar:
            if not isinstance(ay_sonuc, dict):
                continue

            data = [
                ay_sonuc.get('ay_adi', ''),
                format_para(ay_sonuc.get('hesap_brut', 0) or ay_sonuc.get('tam_brut', 0)),
                format_para(ay_sonuc.get('sgk_personel', 0)),
                format_para(ay_sonuc.get('issizlik_personel', 0)),
                format_para(ay_sonuc.get('odenecek_gv', 0)),
                format_para(ay_sonuc.get('toplam_kesinti', 0)),
                format_para(ay_sonuc.get('net_maas', 0)),
                format_para(ay_sonuc.get('isveren_maliyeti', 0)),
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1

    row += 1
    ws[f'A{row}'] = "YILLIK ÖZET"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    ws.cell(row=row, column=1, value="Kalem").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Tutar").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    ozet = sonuc.get('yillik_ozet', {}) if isinstance(sonuc, dict) else {}

    ozet_satirlari = [
        ('Toplam Brüt', ozet.get('toplam_brut', 0)),
        ('Toplam Net', ozet.get('toplam_net', 0)),
        ('Toplam SGK (İşçi)', ozet.get('toplam_sgk_personel', 0)),
        ('Toplam İşsizlik (İşçi)', ozet.get('toplam_issizlik_personel', 0)),
        ('Toplam BES', ozet.get('toplam_bes', 0)),
        ('Toplam Gelir Vergisi', ozet.get('toplam_gv', 0)),
        ('Toplam Damga Vergisi', ozet.get('toplam_dv', 0)),
        ('Toplam Kesinti', ozet.get('toplam_kesinti', 0)),
        ('Toplam Hazine İndirimi', ozet.get('toplam_hazine_indirimi', 0)),
        ('Toplam İşveren Maliyeti', ozet.get('toplam_isveren_maliyeti', 0)),
        ('Ortalama Net Ücret', ozet.get('ortalama_net', 0)),
        ('Ortalama Brüt Ücret', ozet.get('ortalama_brut', 0)),
    ]

    for label, val in ozet_satirlari:
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=format_para(val)).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        if label in ['Toplam Net', 'Toplam İşveren Maliyeti']:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def create_tazminat_excel(sonuc, calisan_adi=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tazminat"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    section_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "KIDEM VE İHBAR TAZMİNATI HESAPLAMASI"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1

    if calisan_adi:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Çalışan: {calisan_adi}"
        row += 1

    row += 1
    hizmet = sonuc.get('hizmet_suresi', {}) if isinstance(sonuc, dict) else {}
    ws[f'A{row}'] = "HİZMET SÜRESİ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    hizmet_data = [
        ('Giriş Tarihi', hizmet.get('giris_tarihi', '-')),
        ('Çıkış Tarihi', hizmet.get('cikis_tarihi', '-')),
        ('Brüt Gün', hizmet.get('brut_gun', 0)),
        ('Kıdem Dışı Gün', hizmet.get('kidem_disi_gun', 0)),
        ('Net Gün', hizmet.get('net_gun', 0)),
        ('Net Hizmet Süresi', hizmet.get('metin', '-')),
    ]

    for label, val in hizmet_data:
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=str(val)).border = thin_border
        row += 1

    row += 1
    ucretler = sonuc.get('ucretler', {}) if isinstance(sonuc, dict) else {}
    ws[f'A{row}'] = "ÜCRET BİLGİLERİ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ucret_data = [
        ('Aylık Brüt Ücret', format_para(ucretler.get('aylik_brut', 0))),
        ('Aylık Ek Ücret', format_para(ucretler.get('aylik_ek', 0))),
        ('Yıllık İkramiye', format_para(ucretler.get('yillik_ikramiye', 0))),
        ('Giydirilmiş Brüt', format_para(ucretler.get('giydirilmis_brut', 0))),
        ('Günlük Brüt', format_para(ucretler.get('gunluk_brut', 0))),
        ('Kıdem Tavanı', format_para(ucretler.get('kidem_tavani', 0))),
        ('Kıdem Matrahı', format_para(ucretler.get('kidem_matrahi', 0))),
    ]

    for label, val in ucret_data:
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=val).border = thin_border
        row += 1

    row += 1
    kidem = sonuc.get('kidem', {}) if isinstance(sonuc, dict) else {}
    ws[f'A{row}'] = "KIDEM TAZMİNATI"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    for col, header in enumerate(['', 'Adet', 'Birim Tutar', 'Tutar'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    if kidem.get('hak_edildi', False):
        kidem_yil = kidem.get('yil', {}) if isinstance(kidem.get('yil'), dict) else {}
        kidem_ay = kidem.get('ay', {}) if isinstance(kidem.get('ay'), dict) else {}
        kidem_gun = kidem.get('gun', {}) if isinstance(kidem.get('gun'), dict) else {}

        kidem_rows = [
            ('Yıl', kidem_yil.get('adet', 0), format_para(kidem_yil.get('birim_tutar', 0)),
             format_para(kidem_yil.get('tutar', 0))),
            ('Ay', kidem_ay.get('adet', 0), format_para(kidem_ay.get('birim_tutar', 0)),
             format_para(kidem_ay.get('tutar', 0))),
            ('Gün', kidem_gun.get('adet', 0), format_para(kidem_gun.get('birim_tutar', 0)),
             format_para(kidem_gun.get('tutar', 0))),
        ]

        for label, adet, birim, tutar in kidem_rows:
            ws.cell(row=row, column=1, value=label).border = thin_border
            ws.cell(row=row, column=2, value=adet).border = thin_border
            ws.cell(row=row, column=3, value=birim).border = thin_border
            ws.cell(row=row, column=4, value=tutar).border = thin_border
            row += 1
        ws.cell(row=row, column=1, value="BRÜT KIDEM TAZMİNATI").border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws.cell(row=row, column=2, value="").border = thin_border
        ws.cell(row=row, column=3, value="").border = thin_border
        ws.cell(row=row, column=4, value=format_para(kidem.get('brut_toplam', 0))).border = thin_border
        ws[f'D{row}'].font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Damga Vergisi (-)").border = thin_border
        ws.cell(row=row, column=4, value=format_para(kidem.get('damga_vergisi', 0))).border = thin_border
        row += 1
        ws.cell(row=row, column=1, value="NET KIDEM TAZMİNATI").border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = highlight_fill
        ws.cell(row=row, column=4, value=format_para(kidem.get('net_toplam', 0))).border = thin_border
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].fill = highlight_fill
        row += 1
    else:
        ws.cell(row=row, column=1, value="Hak Edilmedi (< 1 Yıl)").border = thin_border
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    row += 1
    ihbar = sonuc.get('ihbar', {}) if isinstance(sonuc, dict) else {}
    ws[f'A{row}'] = "İHBAR TAZMİNATI"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    if ihbar.get('hesaplansin', True):
        ihbar_data = [
            ('İhbar Süresi (Hafta)', ihbar.get('sure_hafta', 0)),
            ('İhbar Süresi (Gün)', ihbar.get('sure_gun', 0)),
            ('Günlük Brüt Ücret', format_para(ihbar.get('gunluk_brut', 0))),
            ('BRÜT İHBAR TAZMİNATI', format_para(ihbar.get('brut_toplam', 0))),
            ('Gelir Vergisi (-)', format_para(ihbar.get('gelir_vergisi', 0))),
            ('Damga Vergisi (-)', format_para(ihbar.get('damga_vergisi', 0))),
            ('NET İHBAR TAZMİNATI', format_para(ihbar.get('net_toplam', 0))),
        ]

        for label, val in ihbar_data:
            ws.cell(row=row, column=1, value=label).border = thin_border
            ws.cell(row=row, column=2, value=str(val)).border = thin_border
            if 'BRÜT' in label or 'NET' in label:
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            if 'NET' in label:
                ws[f'A{row}'].fill = highlight_fill
                ws[f'B{row}'].fill = highlight_fill
            row += 1
    else:
        ws.cell(row=row, column=1, value="İhbar Tazminatı Hesaplanmadı").border = thin_border
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    row += 1
    toplam = sonuc.get('toplam', {}) if isinstance(sonuc, dict) else {}
    ws[f'A{row}'] = "GENEL TOPLAM"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    toplam_data = [
        ('Brüt Tazminat Toplamı', format_para(toplam.get('brut_tazminat', 0))),
        ('Gelir Vergisi (-)', format_para(toplam.get('gelir_vergisi', 0))),
        ('Damga Vergisi (-)', format_para(toplam.get('damga_vergisi', 0))),
        ('NET TAZMİNAT TOPLAMI', format_para(toplam.get('net_tazminat', 0))),
    ]

    for label, val in toplam_data:
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=val).border = thin_border
        if 'NET' in label:
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'B{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = highlight_fill
            ws[f'B{row}'].fill = highlight_fill
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output